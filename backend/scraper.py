from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from fpdf import FPDF
from datetime import datetime
import os
import json
from urllib.parse import urlparse, parse_qs
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from concurrent.futures import ThreadPoolExecutor, as_completed
from fpdf.enums import XPos, YPos

# Default API key from environment (set in Render dashboard)
DEFAULT_API_KEY = os.environ.get('YOUTUBE_API_KEY', '')


def extract_id_from_url(url: str):
    """Return (type, id) — type is 'playlist' or 'video'."""
    try:
        parsed = urlparse(url.strip())
        query = parse_qs(parsed.query)

        # youtu.be short links
        if parsed.netloc in ('youtu.be',):
            video_id = parsed.path.lstrip('/')
            if video_id:
                return 'video', video_id

        # playlist takes priority over video when both params present
        if 'list' in query:
            return 'playlist', query['list'][0]
        if 'v' in query:
            return 'video', query['v'][0]

        return None, None
    except Exception as e:
        print(f"Error extracting ID from URL: {e}")
        return None, None


def _build_youtube(api_key: str):
    return build('youtube', 'v3', developerKey=api_key, cache_discovery=False)


def fetch_channel_details(channel_id: str, api_key: str):
    youtube = _build_youtube(api_key)
    response = youtube.channels().list(part='statistics', id=channel_id).execute()
    if not response.get('items'):
        return {'subscribers': 'N/A'}
    stats = response['items'][0].get('statistics', {})
    return {'subscribers': stats.get('subscriberCount', 'N/A')}


def fetch_video_details(video_id: str, api_key: str):
    """Fetch full details for a single video. Raises on API error."""
    youtube = _build_youtube(api_key)
    response = youtube.videos().list(part='snippet,statistics', id=video_id).execute()

    if not response.get('items'):
        return None

    video = response['items'][0]
    snippet = video['snippet']
    stats = video.get('statistics', {})
    channel_id = snippet['channelId']

    try:
        publish_time = datetime.strptime(snippet['publishedAt'], '%Y-%m-%dT%H:%M:%SZ')
        formatted_date = publish_time.strftime('%Y-%m-%d')
    except Exception:
        formatted_date = snippet.get('publishedAt', 'N/A')

    try:
        channel_details = fetch_channel_details(channel_id, api_key)
    except Exception:
        channel_details = {'subscribers': 'N/A'}

    return {
        'Video Title':       snippet.get('title', 'N/A'),
        'Video Link':        f"https://www.youtube.com/watch?v={video_id}",
        'Channel Name':      snippet.get('channelTitle', 'N/A'),
        'Total Subscribers': channel_details['subscribers'],
        'Channel Link':      f"https://www.youtube.com/channel/{channel_id}",
        'Current Views':     stats.get('viewCount', 'N/A'),
        'Video Publish Date': formatted_date,
    }


def fetch_playlist_videos(playlist_id: str, api_key: str):
    """
    Fetch all videos from a playlist.
    Raises HttpError so the caller can detect quota issues.
    """
    youtube = _build_youtube(api_key)
    video_ids = []
    next_page_token = None

    # Step 1 — collect all video IDs from playlist pages
    while True:
        request = youtube.playlistItems().list(
            part='snippet',
            playlistId=playlist_id,
            maxResults=50,
            pageToken=next_page_token,
        )
        response = request.execute()  # raises HttpError on quota/auth failure

        for item in response.get('items', []):
            vid = item['snippet']['resourceId'].get('videoId')
            if vid:
                video_ids.append(vid)

        next_page_token = response.get('nextPageToken')
        if not next_page_token:
            break

    if not video_ids:
        return []

    # Step 2 — fetch details in parallel (max 5 workers)
    videos = []
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {executor.submit(fetch_video_details, vid, api_key): vid for vid in video_ids}
        for future in as_completed(futures):
            try:
                result = future.result()
                if result:
                    videos.append(result)
            except HttpError as e:
                # Re-raise quota errors so main.py can handle them
                raise e
            except Exception as e:
                print(f"Skipping video {futures[future]}: {e}")

    return videos


# ── Export helpers ────────────────────────────────────────────

def save_to_excel(data: list, file_name: str):
    try:
        headers = list(data[0].keys())
        wb = Workbook()
        ws = wb.active
        # Header row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        # Data rows
        for row_idx, item in enumerate(data, 2):
            for col, header in enumerate(headers, 1):
                value = item.get(header, '')
                cell = ws.cell(row=row_idx, column=col, value=value)
                if value and "http" in str(value):
                    cell.font = Font(color="0000FF")
                    cell.hyperlink = str(value)
        wb.save(file_name)
        return len(data)
    except Exception as e:
        print(f"Error saving Excel: {e}")
        raise


def save_to_pdf(data: list, file_name: str):
    try:
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Helvetica", size=12, style='B')
        pdf.cell(0, 10, txt="YouTube Video Details", ln=True, align='C')
        pdf.set_font("Helvetica", size=9)

        for item in data:
            pdf.ln(4)
            for key, value in item.items():
                value = str(value)
                line = f"{key}: {value}"
                if "http" in value:
                    pdf.set_text_color(0, 0, 255)
                    pdf.cell(0, 8, txt=line, link=value, ln=True)
                    pdf.set_text_color(0, 0, 0)
                elif len(line) > 100:
                    pdf.multi_cell(0, 8, txt=line)
                else:
                    pdf.cell(0, 8, txt=line, ln=True)

        pdf.output(file_name)
        return len(data)
    except Exception as e:
        print(f"Error saving PDF: {e}")
        raise


def save_to_json(data: list, file_name: str):
    try:
        with open(file_name, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        return len(data)
    except Exception as e:
        print(f"Error saving JSON: {e}")
        raise
